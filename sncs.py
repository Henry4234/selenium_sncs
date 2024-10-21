##autherise by Henry Tsai
import sys,os, glob,zipfile,re,csv,threading, drivertester
import pandas as pd
import tkinter as tk
from tkinter import ttk,filedialog,messagebox
from tkinter.ttk import Progressbar
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tqdm import tqdm
from time import sleep
from openpyxl import Workbook
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font,Alignment,Border,Side,numbers,PatternFill
from openpyxl.formatting.rule import Rule,CellIsRule,FormulaRule


def worksheet_name_generate(file_path):
    match = re.search(r"\((.*?)\).*_(\d+)\.csv", file_path)
    if match:
        lev_element = match.group(1)
        number_element = match.group(2)
        if number_element=="43052":
            new_string = f"XN-1({number_element})-{lev_element}"
        else:
            new_string = f"XN-2({number_element})-{lev_element}"
        return new_string

    return None

class DOWNLOAD:
    def __init__(self):
        # self.master = master
        # self.progress_window = progress_window
        # self.master.withdraw()
        self.folder_path = filedialog.askdirectory()
        self.pbarper = ["從SNCS爬取所需批號","處理解壓縮檔案","建立整合excel檔","將CSV檔資料放入工作表中","繪製表單","加入所需公式","計算mean值及SD","加入條件格式設定","加上批號訊息及統計時間"]
        # if self.folder_path:
        #     self.download_dir = self.folder_path.replace("/","\\")
            # threading.Thread(target=self.prograss_bar).start()
            # self.progress_window = ProgressWindow(self.master)
            # self.start_download()
            # self.master.after(0, self.start_download)
        #     self.ddd()
        # else:
        #     self.progress_window.close()
            # self.master.destroy()
            # ProgressWindow.close()
    # def start_download(self):
    #     self.master.deiconify()
    #     threading.Thread(target=self.ddd).start()
    #     # self.progress_window.top.mainloop()
    #     # self.ddd()
    
    def outter_border(self,worksheet,s_column, s_index, e_column , e_index,border_style):
        L_border = Border(left=Side(style=border_style),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
        R_border = Border(left=Side(style="thin"),right=Side(style=border_style),top=Side(style="thin"),bottom=Side(style="thin"))
        T_border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style=border_style),bottom=Side(style="thin"))
        B_border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style=border_style))
        TL_border = Border(left=Side(style=border_style),right=Side(style="thin"),top=Side(style=border_style),bottom=Side(style="thin"))
        TR_border = Border(left=Side(style="thin"),right=Side(style=border_style),top=Side(style=border_style),bottom=Side(style="thin"))
        BL_border = Border(left=Side(style=border_style),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style=border_style))
        BR_border = Border(left=Side(style="thin"),right=Side(style=border_style),top=Side(style="thin"),bottom=Side(style=border_style))
        # 轉換欄位字母為對應的索引數字
        s_column_index = column_index_from_string(s_column)
        e_column_index = column_index_from_string(e_column)
        for row in range(s_index, e_index + 1):
            # 設定左邊框
            worksheet.cell(row=row, column=s_column_index).border = L_border
            # 設定右邊框
            worksheet.cell(row=row, column=e_column_index).border = R_border
        for col in range(s_column_index, e_column_index + 1):
            # 設定上邊框
            worksheet.cell(row=s_index, column=col).border = T_border
            # 設定下邊框
            worksheet.cell(row=e_index, column=col).border = B_border
        ##修正左上左下游上右下四角問題
        worksheet.cell(row=s_index, column=s_column_index).border = TL_border
        worksheet.cell(row=s_index, column=e_column_index).border = TR_border
        worksheet.cell(row=e_index, column=s_column_index).border = BL_border
        worksheet.cell(row=e_index, column=e_column_index).border = BR_border

    
    def ddd(self,step):
        # """使用 match-case 模式匹配來執行每個步驟的具體邏輯"""
        # 設定下載路徑
        # self.self.download_dir = self.folder_path
        # self.download_dir = "E:\土城長庚醫院\PanelA"
        match step:
            case "從SNCS爬取所需批號":
                print("正在從SNCS爬取所需批號...")
                if not os.path.exists(self.download_dir):
                    os.makedirs(self.download_dir)


                # self.download_dir = "E:\土城長庚醫院\sncs_download"
                # # 檢查並下載（如果需要） ChromeDriver
                drivertester.check_and_setup_driver()
                # options = webdriver.ChromeOptions()

                # 設定下載路徑和開啟自動下載
                # options.add_experimental_option("prefs", {
                # "download.default_directory": self.download_dir,
                # "download.prompt_for_download": False,
                # "download.directory_upgrade": True,
                # "safebrowsing.enabled": False
                # })
                # options.add_argument("--disable-popup-blocking")
                # options.add_argument("--disable-notifications")
                ##新webdriver至pyinstaller的必要步驟
                # def getDriver():
                #     if getattr(sys,"frozen",False):
                #         webdriver_chrome_path = os.path.join(sys._MEIPASS,"chromedriver.exe")
                #         service = Service(executable_path=webdriver_chrome_path)
                #         driver = webdriver.Chrome(webdriver_chrome_path,service=service,chrome_options=options)
                #     else:
                #         webdriver_chrome_path = './chromedriver'
                #         service = Service(executable_path=webdriver_chrome_path)
                #         driver = webdriver.Chrome(service=service,chrome_options=options)
                #     return driver
                #get web address
                driver = drivertester.get_driver(download_dir=self.download_dir)
                # driver = getDriver()
                #get web address
                driver.get("https://sncs-web.com/quality/login")
                #setting waiting index
                wait = WebDriverWait(driver, 5)
                #key in account & pw
                id = wait.until(EC.element_to_be_clickable((By.ID, "mat-input-0")))
                id.send_keys("sncstw00241")

                pw = wait.until(EC.element_to_be_clickable((By.ID, "mat-input-1")))
                pw.send_keys("huKl00#")
                # loginbtn = driver.find_element(By.XPATH, "/html/body/div[7]/fqc-login/div/div/div/fqc-login-form/div/div[1]/button")
                #click login btn
                loginbtn = driver.find_element(By.CLASS_NAME, "fqc-btn")
                loginbtn.click()
                sleep(2)
                #get to the csv website
                driver.get("https://sncs-web.com/quality/measurement-csv")
                WebDriverWait(driver,10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME,"mat-radio-label-content"))
                )
                #click the cookie policy
                cc = driver.find_element(By.CLASS_NAME, "cc-allow")
                cc.click()
                WebDriverWait(driver,10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME,"mat-radio-label-content"))
                )
                #預設是month，點選成Lot
                lotdot = driver.find_element(By.XPATH, "/html/body/div[7]/m-pages/div/div/div/fqc-intraday-data-csv/fqc-portlet/div/div[2]/div[1]/fqc-search-condition/div/form/div[1]/div/mat-radio-group/mat-radio-button[2]/label/span[1]/span[2]")
                lotdot.click()
                #等待網頁更改成Lot
                WebDriverWait(driver,10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME,"ng-tns-c74-4"))
                )
                # 點選control 的選項"XN-CHECK_CLOSED"
                control_slc = driver.find_element(By.CLASS_NAME, "ng-tns-c74-4")
                control_slc.click()
                # sleep(2)
                control_opt = driver.find_element(By.CLASS_NAME, "mat-option-text")
                control_opt.click()
                # sleep(2)
                ##將需要下載的Lot列成一個List，跑LOOP三次
                lot_lst=["mat-option-18","mat-option-19","mat-option-20"]
                ##Lot_name_lst是之後需要修改rar或者CSV檔名需要用到的
                self.lot_name_lst=[]
                for id in lot_lst:
                    #點選Lot選項，叫出下拉式選單
                    lotno_slc = driver.find_element(By.CLASS_NAME, "ng-tns-c74-6")
                    lotno_slc.click()
                    sleep(2)
                    ##點選相應的Lot No
                    lotno_opt = driver.find_element(By.ID, id)
                    lotno_name = driver.find_element(By.ID, id).find_element(By.CLASS_NAME,"mat-option-text").text
                    self.lot_name_lst.append(lotno_name)
                    lotno_opt.click()
                    # sleep(2)
                    #點選搜尋按鈕
                    src_btn = driver.find_element(By.XPATH, "/html/body/div[7]/m-pages/div/div/div/fqc-intraday-data-csv/fqc-portlet/div/div[2]/div[1]/fqc-search-condition/div/form/div[2]/div[4]/button")
                    src_btn.click()
                    ##等待網頁下方出現chkbox，再繼續操作
                    WebDriverWait(driver,30).until(
                        EC.presence_of_all_elements_located((By.CLASS_NAME,"mat-row"))
                    )
                    # sleep(5)
                    spinner_locator = (By.XPATH, '/html/body/div[7]/m-pages/div/div/div/fqc-intraday-data-csv/fqc-portlet/div/div[2]/div[2]/fqc-analyzer-list/div[1]/mat-spinner')
                    WebDriverWait(driver, 30).until(EC.invisibility_of_element_located(spinner_locator))
                    #點選兩個chkbox
                    chkbox_1 = driver.find_element(By.XPATH, "/html/body/div[7]/m-pages/div/div/div/fqc-intraday-data-csv/fqc-portlet/div/div[2]/div[2]/fqc-analyzer-list/div[3]/div/table/tbody/tr[1]/td[1]/mat-checkbox")
                    chkbox_1.click()

                    chkbox_2 = driver.find_element(By.XPATH, "/html/body/div[7]/m-pages/div/div/div/fqc-intraday-data-csv/fqc-portlet/div/div[2]/div[2]/fqc-analyzer-list/div[3]/div/table/tbody/tr[2]/td[1]/mat-checkbox")
                    chkbox_2.click()

                    # WebDriverWait(driver,10).until(
                    #   EC.presence_of_element_located((By.CLASS_NAME,"cdk-overlay-6"))
                    # )
                    sleep(3)
                    ##點選下載按鈕
                    download_btn = driver.find_element(By.CLASS_NAME, "mat-bottom-sheet-container").find_element(By.TAG_NAME,"button")
                    download_btn.click()
                    sleep(5)
                driver.quit()
                # print(self.lot_name_lst)
                # self.progress_window.update_progress(20,"處理解壓縮檔案")
                
            case "處理解壓縮檔案":
                print("正在解壓縮檔案...")
                ######更改檔名，改完之後解壓縮
                files = glob.glob(os.path.join(self.download_dir, '*'))
                sorted_files = sorted(files)
                ##sorted_files會以list型式存在，裡面包tuple
                # [('#3100(L3)_CL   2023-07-02', '.zip'), ('#3101(L1)_CL   2023-07-02', '.zip'), ('#3101(L2)_CL   2023-07-02', '.zip')]
                # print(sorted_files)
                ##利用迴圈將資料夾中所有zip檔解壓縮
                # file_names = [os.path.splitext(os.path.basename(file))[0] for file in sorted_files]

                ##重新命名zip檔
                new_file_path_list=[]
                for old_file, new_file in zip(sorted_files, self.lot_name_lst):
                    new_file_path = os.path.join(self.download_dir, new_file + os.path.splitext(old_file)[1])
                    os.rename(old_file, new_file_path)
                    new_file_path_list.append(new_file_path)

                ##更新資料名稱
                sorted_files = sorted(files)
                ##利用迴圈將資料夾中所有zip檔解壓縮
                for zip_file in new_file_path_list:
                    zip_path = zip_file
                    folder_name = os.path.splitext(zip_file)[0]
                    # folder_path_1 = os.path.join(self.download_dir, folder_name)
                    #利用壓縮檔名作為資料夾檔名儲存
                    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                        zip_ref.extractall(folder_name)
                ##所有資料夾名稱
                folder = os.listdir(self.download_dir)
                #self.folder_names = 所有子資料夾的名稱
                self.folder_names = [item for item in folder if os.path.isdir(os.path.join(self.download_dir, item))]
                # print(self.folder_names)
                # self.progress_window.update_progress(35,"建立整合excel檔")
            case"建立整合excel檔":
                print("正在建立Excel文件...")
                ######建立一個excel檔，將解壓縮後的CSV檔放入excel檔中
                self.wb = Workbook()
                # self.progress_window.update_progress(50,"將CSV檔資料放入工作表中")
            case "將CSV檔資料放入工作表中":
                print("正在將CSV數據寫入工作表...")
                for name in self.folder_names:
                    folder_path_2 = os.path.join(self.download_dir, name)
                    
                    for file_name in os.listdir(folder_path_2):
                        file_path_3 = os.path.join(folder_path_2, file_name)
                        ws_name = worksheet_name_generate(file_path_3)
                        # print(ws_name)
                        self.ws = self.wb.create_sheet(title = ws_name)
                        df_1 = pd.read_csv(file_path_3)
                        for r in dataframe_to_rows(df_1, index=False, header=True):
                            self.ws.append(r)
                        # with open(file_path_3, 'r') as file:
                        #     csv_reader = csv.reader(file)
                        #     for row in csv_reader:
                        #         self.ws.append(row)
                ##設置第一個工作表"worksheet"，建立所需表單
                # self.progress_window.update_progress(65,"繪製表單")
            case "繪製表單":
                print("正在繪製表單...")
                self.ws = self.wb["Sheet"]
                ##外邊框設置
                #ex: A3:D5 要外粗邊框
                # A3:A5的left style = "medium"
                # D3:D5的right style = "medium"
                # A3:D3的top style = "medium"
                # A5:D5的bottom style = "medium"
                

                self.title_f = Font(name='微軟正黑體', size=18)  #標題字體設定
                self.subtitle_f = Font(name='微軟正黑體', size=14)  #副標題字體設定
                self.context_f = Font(name='微軟正黑體', size=12)  #內文字體設定


                self.ws["A2"].value = "Online SNCS worldwidemean"
                self.ws["A2"].font = self.title_f
                self.ws.merge_cells("A2:E2")
                self.ws["A3"].value = "統計時間:"
                self.ws["A3"].font = self.title_f
                self.ws.merge_cells("A3:B3")
                self.ws["A3"].alignment = Alignment(horizontal='right', vertical='center')
                ##[A4:D4]統計時間

                ##[D6,K6,R6]Lot No.

                ##有merge的
                merge_ranges = [
                    ['A5:A8', '序號'],['A28:A29','序號'],
                    ['B5:B8', '檢驗項目'],['B28:B29','檢驗項目'],['M28:M29','項目'],
                    ['C5:I5', 'Level 1'],['C28:E28', 'Level 1'],['N28:O28', 'Level 1'],
                    ['J5:P5', 'Level 2'],['F28:H28', 'Level 2'],['P28:Q28', 'Level 2'],
                    ['Q5:W5', 'Level 3'],['I28:K28', 'Level 3'],['R28:S28', 'Level 3'],
                    ['C7:D7', 'SNCS'],['J7:K7', 'SNCS'],['Q7:R7', 'SNCS'],
                    ['E7:E8', '現行SD'],['L7:L8', '現行SD'],['S7:S8', '現行SD'],
                    ['F7:G7', 'XN-1'],['M7:N7', 'XN-1'],['T7:U7', 'XN-1'],
                    ['H7:I7', 'XN-2'],['O7:P7', 'XN-2'],['V7:W7', 'XN-2'],
                    ['M27:S27','CHECK儀器和SNCS有無偏移 (標準：NRBC-±10%，OTHER-±3%，超過字會反紅)'],
                    ["D6:I6",""],["K6:P6",""],["R6:W6",""],
                    ["M47:S47","※若超過上述標準，請先確認是否偏移1SD以上，若有，請跟組長討論或請工程師來調整"],["M48:S48","※標準為暫定，可再與組長討論調整"],

                ]
                #單格
                single_cell=[
                    ['C6','Lot:'],['J6','Lot:'],['Q6','Lot:'],
                    ['C8', 'mean'],['F8', 'mean'],['H8', 'mean'],
                    ['J8', 'mean'],['M8', 'mean'],['O8', 'mean'],
                    ['Q8', 'mean'],['T8', 'mean'],['V8', 'mean'],
                    ['D8', 'SD'],['G8', 'SD'],['I8', 'SD'],
                    ['K8', 'SD'],['N8', 'SD'],['P8', 'SD'],
                    ['R8', 'SD'],['U8', 'SD'],['W8', 'SD'],
                    ['C29','自求'],['D29','SNCS'],['E29','SDI'],
                    ['F29','自求'],['G29','SNCS'],['H29','SDI'],
                    ['I29','自求'],['J29','SNCS'],['K29','SDI'],
                    ['N29','XA'],['O29','XB'],
                    ['P29','XA'],['Q29','XB'],
                    ['R29','XA'],['S29','XB'],

                ]
                #檢驗項目內容
                self.items = ["RBC","HGB","HCT","MCV","MCH","MCHC","RDW-SD","RDW-CV","PLT","WBC","BASO%","NEUT%","LYMPH%","MONO%","EO%","NRBC%","RET%",]
                standard_SD = [
                    ["0.03","0.07","0.30","1.00","0.40","0.50","0.50","0.20","7.00","0.10","0.20","1.20","1.50","1.50","0.80","2.00","0.30"],
                    ["0.04","0.14","0.50","0.80","0.40","0.50","0.50","0.20","10.00","0.16","0.20","1.10","1.20","1.10","0.80","2.00","0.20"],
                    ["0.03","0.07","0.30","1.00","0.40","0.50","0.50","0.20","7.00","0.10","0.20","1.20","1.50","1.50","0.80","2.00","0.30"]
                    ]
                #合併单元格
                for merge_range, _ in merge_ranges:
                    self.ws.merge_cells(merge_range)
                # 设置合并单元格的对齐方式和值
                for merge_range, value in merge_ranges:
                    cell_range = self.ws[merge_range]
                    top_left_cell = cell_range[0][0]
                    top_left_cell.alignment = Alignment(horizontal='center', vertical='center')
                    top_left_cell.value = value
                    top_left_cell.font = self.subtitle_f
                #單一儲存格
                for cell, value in single_cell:
                    self.ws[cell].value = value
                    self.ws[cell].alignment = Alignment(horizontal='center', vertical='center')
                    self.ws[cell].font = self.subtitle_f
                #序號值 & 項目值
                for i in range(9,26):
                    dig = self.ws["A" + str(i)]
                    item = self.ws["B" + str(i)]
                    dig.value = i - 8
                    dig.font = self.context_f
                    item.value = self.items[i - 9]
                    item.font = self.context_f
                #放入固定SD
                for i in range(5,25,7):
                    for j in range(9,26):
                        cell = self.ws[get_column_letter(i) + str(j)]
                        # print(cell)
                        cell.value = float(standard_SD[i//7][j-9])
                        cell.font = self.context_f
                ##設置框線
                border_medium = Border(
                    left=Side(style="medium"),
                    right=Side(style="medium"),
                    top=Side(style="medium"),
                    bottom=Side(style="medium")
                )
                self.border_thin = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
                border_double = Border(
                    left=Side(style="double"),
                    right=Side(style="double"),
                    top=Side(style="double"),
                    bottom=Side(style="double")
                )
                for row in self.ws["A5:W25"]:
                    for cell in row:
                        cell.border = self.border_thin
                self.outter_border(worksheet=self.ws,s_column='A', s_index=5, e_column='W' , e_index=8,border_style="medium")
                self.outter_border(worksheet=self.ws,s_column='A', s_index=9, e_column='B' , e_index=25,border_style="medium")
                self.outter_border(worksheet=self.ws,s_column='C', s_index=9, e_column='I' , e_index=25,border_style="medium")
                self.outter_border(worksheet=self.ws,s_column='J', s_index=9, e_column='P' , e_index=25,border_style="medium")
                self.outter_border(worksheet=self.ws,s_column='Q', s_index=9, e_column='W' , e_index=25,border_style="medium")
                #小數點一位
                self.ws["B26"].value = "※ mean小數點依據：現行發報告位數多一位。"
                self.ws["B26"].font = self.context_f
                self.ws.merge_cells("B26:F26")

                #####自求mean & SDI
                self.ws["A27"].value = "*預計設定之mean(XN-1 & XN-2 平均)"
                self.ws["A27"].font = self.title_f
                self.ws.merge_cells("A27:F27")
                #序號值 & 項目值
                for i in range(30,47):
                    dig = self.ws["A" + str(i)]
                    item = self.ws["B" + str(i)]
                    dig.value = i - 29
                    dig.font = self.context_f
                    item.value = self.items[i - 30]
                    item.font = self.context_f
                #(M27:U27)放入項目
                for i in range(30,47):
                    cell = self.ws['M' + str(i)]
                    cell.value = self.items[i - 30]
                    cell.font = self.context_f
                ##加入公式
                #C30 = (A9 + H9) / 2
                #C31 = (A10 + H10) / 2
                #exception:C47 = H25
                # self.progress_window.update_progress(75,"加入所需公式")
            case "加入所需公式":
                print("正在加入所需公式...")
                i=0
                for col in range(3,12,3):
                    for j in range(30,47):
                        cell_1 = self.ws[get_column_letter(col) + str(j)]
                        cell_2 = self.ws[get_column_letter(col + 1) + str(j)]
                        cell_3 = self.ws[get_column_letter(col + 2) + str(j)]
                        col_1 = get_column_letter(col+3+(4*i))
                        col_2 = get_column_letter(col+5+(4*i))
                        #特殊的公式(RBC應該要求到小數點後三位/PLT應該要求到小數點後一位)
                        if j==30 or j==38:
                            match j:
                                case 30:
                                    formula_1 = "=ROUND((%s%d + %s%d)/2,3)"%(col_1,j-21,col_2,j-21)
                                    formula_2 = "=ROUND(%s%d,3)"%(get_column_letter(col+(4*i)),j-21)
                                case 38:  
                                    formula_1 = "=ROUND((%s%d + %s%d)/2,1)"%(col_1,j-21,col_2,j-21)
                                    formula_2 = "=ROUND(%s%d,1)"%(get_column_letter(col+(4*i)),j-21)
                        else:    
                            formula_1 = "=ROUND((%s%d + %s%d)/2,2)"%(col_1,j-21,col_2,j-21)
                            formula_2 = "=ROUND(%s%d,2)"%(get_column_letter(col+(4*i)),j-21)
                        formula_3 = "=(%s%d - %s%d)/%s%d"%(get_column_letter(col),j,get_column_letter(col+1),j,get_column_letter(col+2+(4*i)),j-21)
                        cell_1.value = formula_1
                        cell_2.value = formula_2
                        cell_3.value = formula_3
                        cell_1.font = self.context_f
                        cell_2.font = self.context_f
                        cell_3.font = self.context_f
                        cell_3.number_format = numbers.FORMAT_PERCENTAGE
                    self.ws[get_column_letter(col) + str(46)].value = "=ROUND(%s25,2)"%(get_column_letter(col+5+(4*i)))
                    i+=1

                RBCPLT_lst=[30,38]
                for colno in RBCPLT_lst:
                    formula_RBC = "=ROUND((%s%d + %s%d)/2,2)"%(col_1,j-21,col_2,j-21)
                    formula_PLT = "=ROUND((%s%d + %s%d)/2,2)"%(col_1,j-21,col_2,j-21)
                ##框線(A28:K46)
                #內框線
                for row in self.ws["A28:K46"]:
                    for cell in row:
                        cell.border = self.border_thin
                #外框線
                self.outter_border(worksheet=self.ws,s_column='A', s_index=28, e_column='K' , e_index=46,border_style="medium")
                ##(M27:U27)CHECK儀器和SNCS有無偏移
                
                #加入外框線與內框線
                #內框線
                for row in self.ws["M27:S46"]:
                    for cell in row:
                        cell.border = self.border_thin
                #外框線
                self.outter_border(worksheet=self.ws,s_column='M', s_index=27, e_column='S' , e_index=46,border_style="medium")

                ###建立好框線開始塞資料

                #在子資料夾中，提取所有文件名稱
                def get_all_filenames(directory):
                    filenames = []
                    for root, dirs, files in os.walk(directory):
                        for file in files:
                            file_path = os.path.join(root, file)
                            filenames.append(file_path)
                    return filenames
                self.path_list=[]
                #add dir
                for child in self.folder_names:
                    new_dir = self.download_dir +"\\"+ child
                    pathes = get_all_filenames(new_dir)
                    for path in pathes:
                        self.path_list.append(path)
                #把六個csv一一打開讀取放入資料
                # self.progress_window.update_progress(85,"計算mean值及SD")
            case "計算mean值及SD":
                print("正在計算mean值及SD...")
                for di in self.path_list:
                    #提取lev
                    lev = di.split("(")
                    lev = lev[1].split(")")[0]
                    # print(lev)
                    #提取機台(43052 or 43056)
                    mac = di.split("_")
                    mac = mac[-1].split(".")[0]
                    # print(mac)
                    #建立dataframe
                    self.df = pd.read_csv(di)
                    # print(df)
                    #利用日期&時間排序
                    df = self.df.sort_values(by=['Date','Time'],ascending=False)
                    #掛filter依序把items裡面的 PG mean & PG SD 找出來
                    PG_mlst,PG_slst,self_mean_lst,self_SD_lst = [],[],[],[]
                    for item in self.items:
                        filter = df['Parameter'] == item
                        sncs_mean = df[filter].head(1)['PG mean'].values
                        sncs_SD = df[filter].head(1)['PG SD'].values
                        self_mean = df[filter]['Data'].mean()
                        self_SD = df[filter]['Data'].std()
                        PG_mlst.append(sncs_mean)
                        PG_slst.append(sncs_SD)
                        self_mean_lst.append(round(self_mean,3))
                        self_SD_lst.append(round(self_SD,3))
                    # print(self_mean_lst)
                    # print(self_SD_lst)
                    #因為43056 Peer gtoup有RET，所以採用XN-2的PG mean & SD
                    if mac=="43056":
                        #chk lev放到固定的位置
                        match lev:
                            case "L1":
                                col_letters = [3,4]
                                meanSD_col =[8,9]
                                XB_col =[15]
                            case "L2":
                                col_letters = [10,11]
                                meanSD_col =[15,16]
                                XB_col =[17]
                            case "L3":
                                col_letters = [17,18]
                                meanSD_col =[22,23]
                                XB_col =[19]
                        for j in range(0,2):
                            for i in range(9,26):
                                cell = get_column_letter(col_letters[j]) + str(i)
                                meanSD_cell = get_column_letter(meanSD_col[j]) + str(i)
                                XB_cell = get_column_letter(XB_col[0]) + str(i+21)
                                if j == 0:
                                    self.ws[cell].value = PG_mlst[i-9][0]
                                    self.ws[meanSD_cell].value = self_mean_lst[i-9]
                                    self.ws[XB_cell].value = (self_mean_lst[i-9]-PG_mlst[i-9][0])/PG_mlst[i-9][0]
                                    self.ws[cell].font = self.context_f
                                    self.ws[meanSD_cell].font = self.context_f
                                    self.ws[XB_cell].font = self.context_f
                                    self.ws[XB_cell].number_format = numbers.FORMAT_PERCENTAGE_00
                                else:
                                    self.ws[cell].value = PG_slst[i-9][0]
                                    self.ws[cell].font = self.context_f
                                    self.ws[cell].number_format = '0.00'
                                    self.ws[meanSD_cell].value = self_SD_lst[i-9]
                                    self.ws[meanSD_cell].font = self.context_f
                    else:
                        match lev:
                            case "L1":
                                meanSD_col =[6,7]
                                XA_col =[14]
                            case "L2":
                                meanSD_col =[13,14]
                                XA_col =[16]
                            case "L3":
                                meanSD_col =[20,21]
                                XA_col =[18]
                        for j in range(0,2):
                            for i in range(9,26):
                                cell = get_column_letter(meanSD_col[j]) + str(i)
                                XA_cell = get_column_letter(XA_col[0]) + str(i+21)
                                if j == 0:
                                    self.ws[cell].value = self_mean_lst[i-9]
                                    self.ws[cell].font = self.context_f
                                    try:
                                        self.ws[XA_cell].value = (self_mean_lst[i-9]-PG_mlst[i-9][0])/PG_mlst[i-9][0]
                                    except IndexError:
                                        pass
                                    self.ws[XA_cell].font = self.context_f
                                    self.ws[XA_cell].number_format = numbers.FORMAT_PERCENTAGE_00
                                else:
                                    self.ws[cell].value = self_SD_lst[i-9]
                                    self.ws[cell].font = self.context_f
                                    self.ws[cell].number_format = '0.00'
                                    
                ###設定多個條件的條件式格式設定
                # self.progress_window.update_progress(90,"加入條件格式設定")
            case "加入條件格式設定":
                print("正在加入條件格式設定...")
                #range_1 & 2為SNCS有無偏移
                ranges_1 = ["N30:S44","N46:S46"]
                ranges_2 = "N45:S45"
                #range_3 為Precision
                ranges_3 = ["N30:S44","N46:S46"]
                ranges_4 = ["G9:G24","N9:N24","U9:U24","I9:I25","P9:P25","W9:W25"]
                # 建立填充和字體的樣式
                fill = PatternFill(start_color="FFD2D2", end_color="FFD2D2", fill_type="solid")
                red_font = Font(color="FF0000")
                # 建立條件式格式設定規則
                rule1 = CellIsRule(operator="greaterThan", formula=[0.03], stopIfTrue=False, fill=fill, font=red_font)
                rule2 = CellIsRule(operator="lessThan", formula=[-0.03], stopIfTrue=False, fill=fill, font=red_font)
                rule3 = CellIsRule(operator="greaterThan", formula=[0.10], stopIfTrue=False, fill=fill, font=red_font)
                rule4 = CellIsRule(operator="lessThan", formula=[-0.10], stopIfTrue=False, fill=fill, font=red_font)

                for set_range in ranges_1:
                    self.ws.conditional_formatting.add(set_range, rule1)
                    self.ws.conditional_formatting.add(set_range, rule2)
                self.ws.conditional_formatting.add(ranges_2, rule3)
                self.ws.conditional_formatting.add(ranges_2, rule4)

                formula=[
                    ['=$G9>1.5*$D9'],
                    ['=$N9>1.5*$K9'],
                    ['=$U9>1.5*$R9'],
                    ['=$I9>1.5*$D9'],
                    ['=$P9>1.5*$K9'],
                    ['=$W9>1.5*$R9']
                ]
                for i in range(0,6):
                    rule5 = FormulaRule(formula=formula[i],fill=fill,font=red_font)
                    self.ws.conditional_formatting.add(ranges_4[i], rule5)
                ##最後加上批號訊息跟統計時間
                #尋找起始時間與結束時間
                # self.progress_window.update_progress(98,"加上批號訊息及統計時間...")
            case "加上批號訊息及統計時間":
                print("正在加上批號訊息及統計時間...")
                df = self.df.sort_values(by=['Date','Time'],ascending=False)
                start_date = df.tail(1)["Date"].values
                end_date = df.head(1)["Date"].values
                #合併成string = analyse_date
                analyse_date = start_date[0] + "-" + end_date[0]
                self.ws["C3"].value = analyse_date
                self.ws["C3"].font = self.subtitle_f
                #合併儲存格，靠左對齊，水平置中
                self.ws.merge_cells("C3:F3")
                self.ws["C3"].alignment = Alignment(horizontal='left', vertical='center')
                #新增批號訊息
                for foldername in self.folder_names:
                    split_lst = foldername.split("(")
                    lot = split_lst[1][0:2]
                    match lot:
                        case "L1":
                            lot_str = split_lst[0] + str(1101)
                            cell = self.ws["D6"]
                        case "L2":
                            lot_str = split_lst[0] + str(1102)
                            cell = self.ws["K6"]
                        case "L3":
                            lot_str = split_lst[0] + str(1103)
                            cell = self.ws["R6"]
                    cell.value = lot_str
                    cell.font = self.subtitle_f
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                ##設定B欄寬度
                self.ws.column_dimensions['B'].width = 14
                ####填滿網底
                #設定三個顏色
                title_fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
                item_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
                mean_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                #設定range(表頭，項目，mean&SD)
                titles_range = ["A5:W8","A28:K29","M28:S29"]
                items_range = ["A9:B25","A30:B46","M30:M46"]
                meanSD_range = ["C9:D25","J9:K25","Q9:R25"]
                for fillrange in titles_range:
                    slcrange = self.ws[fillrange]
                    for row in slcrange:
                        for cell in row:
                            cell.fill = title_fill
                for fillrange in items_range:
                    slcrange = self.ws[fillrange]
                    for row in slcrange:
                        for cell in row:
                            cell.fill = item_fill
                for fillrange in meanSD_range:
                    slcrange = self.ws[fillrange]
                    for row in slcrange:
                        for cell in row:
                            cell.fill = mean_fill
                excel_path = os.path.join(self.download_dir, "merge.xlsx")
                self.wb.save(excel_path)
                tk.messagebox.showinfo('土城長庚醫院檢驗科', message='下載整理成功!檔案在%s'%(excel_path))
                # self.master.destroy()
                # self.progress_window.close()
                return
            case _:
                print(f"未知的步驟: {step}")
    def run(self):
        """控制流程並顯示進度條"""
        if self.folder_path:
            self.download_dir = self.folder_path.replace("/","\\")
            with tqdm(total=len(self.pbarper), desc="處理進度", ncols=100) as pbar:
                for step in self.pbarper:
                    self.ddd(step)  # 執行每一個步驟
                    pbar.set_postfix_str(f"正在處理: {step}")  # 顯示當前步驟
                    pbar.update(1)  # 更新進度條
            print("下載整理成功!")
            return
        else:
            print("byebye!")
            return
    # def prograss_bar(self):
    #     # self.parent = root
    #     self.top = tk.Tk()
    #     # self.top = self.parent
    #     # self.master.deiconify()
    #     self.top.title("進度條")
    #     self.top.geometry("300x150")
        
    #     self.label = tk.Label(self.top, text="從SNCS爬取所需批號")
    #     self.label.pack(pady=10)
        
    #     self.progressbar = ttk.Progressbar(self.top, length=200)
    #     self.progressbar.pack(pady=10)
        
    #     # 創建一個Label顯示目前進度的資訊
    #     self.progress_label = tk.Label(self.top, text="進度：1%")
    #     self.progress_label.pack(pady=10)
    #     # 在另一個執行緒中執行耗時的操作
    #     self.progressbar.start()
    #     threading.Thread(target=DOWNLOAD).start()
    #     self.top.mainloop()
        

    # def update_progress(self, progress,text):
    #     # 更新進度條數值和顯示的進度資訊
    #     self.progressbar['value'] = progress
    #     self.progress_label['text'] = f"進度：{progress}%"
    #     self.label.config(text=text)

    def close(self):
        # oldmaster.deiconify()
        self.root.destroy()


class ProgressWindow:
    def __init__(self,parent):
        super().__init__()
        self.parent = parent
        self.top = self.parent
        self.top.title("進度條")
        self.top.geometry("300x150")
        
        self.label = tk.Label(self.top, text="從SNCS爬取所需批號")
        self.label.pack(pady=10)
        
        self.progressbar = ttk.Progressbar(self.top, length=200)
        self.progressbar.pack(pady=10)
        
        # 創建一個Label顯示目前進度的資訊
        self.progress_label = tk.Label(self.top, text="進度：1%")
        self.progress_label.pack(pady=10)
        # 在另一個執行緒中執行耗時的操作
        self.progressbar.start()
        self.download = DOWNLOAD(master=self.parent,progress_window = self)
        # threading.Thread(target=DOWNLOAD(master=self.parent,progress_window = self))
        # threading.Thread(target=DOWNLOAD(master=self.parent,progress_window = self)).start()
        # self.top.mainloop()
        

    def update_progress(self, progress,text):
        # 更新進度條數值和顯示的進度資訊
        self.progressbar['value'] = progress
        self.progress_label['text'] = f"進度：{progress}%"
        self.label.config(text=text)

    def close(self):
        # oldmaster.deiconify()
        # self.download.stop()
        
        self.top.destroy()
        


def main():
    # root =tk.Tk()
    # ProgressWindow(root)
    D = DOWNLOAD()
    D.run()         # 執行類別中的處理流程
    # D.close()

    # root.mainloop()

if __name__ == '__main__':  
    main()
